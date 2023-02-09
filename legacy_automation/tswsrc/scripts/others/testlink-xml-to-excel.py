"""
Script to convert xml file to xlsx
"""
__author__ = 'manoj'
import os
import copy
import pprint
import logging
import xml.etree.ElementTree as xET
from xml.dom import minidom
import xlsxwriter
from libx.html2text import html2text
from openpyxl import load_workbook
from collections import Iterable
import re

EXPECTED_RESULTS = u'Expected Results'
ROOT_PLACEHOLDER = u'~root'
DELIMITER = u'|'
HEAD_ROWS = 2
ROW_HEIGHT = 15
COL_WIDTH_STEPS = 50

#todo: Add Flags as new column in the excel sheet based on errors detected during export
#todo: Extract details


def prettify(elem):
    """Return a pretty-printed XML string for the Element."""
    rough_string = xET.tostring(elem, 'utf-8')
    parsed = minidom.parseString(rough_string)
    return parsed.toprettyxml(indent="  ")


def hook_format_data_written_to_cell(data_tuple):
    """

    :param data_tuple: Entire row of a work sheet
    :return: list
    """
    data_l = []
    for cell in data_tuple:
        data_l.append(cell.replace('\n', '<br>\n'))
    return data_l


class TestLinkXMLUtils:
    """Helps to extract data from testlink xml files"""

    def __init__(self):
        #### instance variables
        self.worksheet_dict = {}
        self.suites = {}
        self.xml_root = None
        self.suite_names = None

    @staticmethod
    def _assert_element_type(node):
        """
        Asserts Whether the node is instance of xml.etree.ElementTree.Element or not.
        :param node: Element
        :return: Boolean
        :raises: TypeError
        """
        if isinstance(node, xET.Element):
            return True
        raise TypeError('suite_node must be Element. found %s' % type(node))

    @staticmethod
    def extract_tests(suite_node, recurisve):
        """
        Hook to process the test suite_node in XML to a desired flat file structure

        :param recurisve: True will process all sub suites recursively
        :param suite_node: Element suite_node of test suite
        """
        TestLinkXMLUtils._assert_element_type(suite_node)
        if recurisve:
            return [suite for suite in suite_node.iter('testcase')]
        return suite_node.findall('testcase')

    @staticmethod
    def process_testcase(suite_node, parent_prefix=''):
        """
        Hook to process the test suite_node in XML to a desired flat file structure

        :param suite_node: Element suite_node of test suite
        """
        TestLinkXMLUtils._assert_element_type(suite_node)

        tc_dict = {'name': suite_node.attrib['name'], 'internalid': suite_node.attrib['internalid']}
        #### Extract test case suite_node and form a dictionary (Intermediate format!)

        for tag_name in ['node_order', 'externalid', 'summary', 'version',
                         'preconditions', 'execution_type', 'importance']:
            tc_dict[tag_name] = (suite_node.find(tag_name)).text

        steps_l = []
        for step_e in suite_node.iter('step'):
            step = {}
            for tag in ['step_number', 'actions', 'expectedresults', 'execution_type']:
                step[tag] = (step_e.find(tag)).text
            steps_l.append(step)
        tc_dict['steps'] = steps_l

        custom_d = {}
        custom_l = []
        for custom_e in suite_node.iter('custom_field'):
            for tag in ['name', 'value']:
                custom_d[tag] = (custom_e.find(tag)).text
                custom_l.append(custom_d)
        tc_dict['custom_field'] = custom_l

        # pprint.pprint(tc_dict)
        if parent_prefix:
            tc_dict['name'] = parent_prefix.rpartition('|')[-1] + ": " + tc_dict['name']
        return tc_dict

    @staticmethod
    def extract_suite_data(suite_node):
        """
        :param suite_node: Element tree node for a given suite
        """
        if not isinstance(suite_node, xET.Element):
            raise TypeError('suite_node must be Element. found %s' % type(suite_node))
        for testcase in suite_node.iter('testcase'):
            print(testcase.attrib)
            TestLinkXMLUtils.process_testcase(testcase)

    @staticmethod
    def get_suites(current_node, max_suite_level, current_depth=0, parent_prefix="", suite_list=None,
                   delimiter=DELIMITER, errs=None):
        """
        Recursive function to extract xml data in to a dict

        :param current_node: instance of Element object that will be taken for processing
        :param max_suite_level: The maximum level of nodes that will be processed
        :param current_depth: Denotes the current depth in the search
        :param parent_prefix:
        :param suite_list:  List of suites for recursive call
        :param delimiter: Delimiter for work sheet names
        :param errs:
        :return: suite_list, errs
        """
        if suite_list is None:
            suite_list = []
        if errs is None:
            errs = []
        if not isinstance(current_node, xET.Element):
            raise TypeError('current_node must be Element. found %s' % type(current_node))

        if not parent_prefix:
            current_node_name = current_node.attrib['name']
        else:
            current_node_name = parent_prefix + delimiter + current_node.attrib['name']

        children_names = [child_node.attrib['name'].lower() for child_node in current_node.findall('testsuite')]
        details_text = ''
        try:
            details_text = (current_node.find('details')).text
        except AttributeError:
            pass

        if len(children_names) == 0:
            leaf_bool = True
        else:
            leaf_bool = False

        #append here
        suite_list.append({'name': current_node_name,
                           'node': current_node,
                           'leaf': leaf_bool,
                           'info': details_text})

        if current_depth >= max_suite_level:
            return suite_list
        current_depth += 1

        # check for errors
        for child_node in current_node.findall('testsuite'):
            if children_names.count(child_node.attrib['name'].lower()) > 1:
                errs.append('Suite names are case insensitive and unique. Error for: %s' % child_node.attrib['name'])
            child_node_name = current_node_name+delimiter+child_node.attrib['name']
            if len(child_node_name) > 31:
                errs.append('Suite names must be <= 31 chars. Error for: %s' % child_node_name)

            TestLinkXMLUtils.get_suites(child_node, max_suite_level,
                                        current_depth=current_depth,
                                        parent_prefix=current_node_name,
                                        suite_list=suite_list,
                                        errs=errs)

        return suite_list, errs
            # find duplicates

    @staticmethod
    def construct_testsuite_tree(current_root_node, worksheets, worksheets_dict=None, prefix="", delimiter=DELIMITER):
            #input validation!
        if worksheets_dict is None:
            worksheets_dict = {}
        if not isinstance(worksheets, list):
            raise TypeError('worksheets must be a list! found: %s' % type(worksheets))
        if not isinstance(worksheets_dict, dict):
            raise TypeError('worksheets_dict must be a dict! found: %s' % type(worksheets_dict))

            #initialization
        if prefix:
            children = [sheet for sheet in worksheets if sheet.startswith(prefix)]
        else:
            if not worksheets[0] == ROOT_PLACEHOLDER:
                raise ValueError('worksheets[0] must be "~roots". Found: %s' % worksheets[0])
            worksheets.pop(0)
            children = [sheet for sheet in worksheets if DELIMITER not in sheet]


        for child in children:
            worksheets_dict[child] = {}
        print(children)
        print(worksheets)
        print(worksheets_dict)


class TestLinkXMLExcel2Xml(TestLinkXMLUtils):
    """
    Converted a valid, pre formatted excel file into a test link importable xml file
    """
    def __init__(self, excel_file):
        #### instance variables
        """

        :param excel_file: Input excel file
        """
        TestLinkXMLUtils.__init__(self)

        if not os.path.isfile(excel_file):
            raise IOError('The excel file does not exist: %s' % excel_file)
            #get the list of worksheets
        workbook = load_workbook(excel_file, use_iterators=True)
        worksheets = workbook.get_sheet_names()

            #construct the xml tree
        xml_root = xET.Element("testsuite")
        node_order_node = xET.SubElement(xml_root, 'node_order')
        node_order_node.text = '0'
        details_node = xET.SubElement(xml_root, 'details')
        details_node.text = ' '

        TestLinkXMLUtils.construct_testsuite_tree(xml_root, worksheets)
        print(prettify(xml_root))
        pprint.pprint(xET.dump(xml_root))


    def tests_to_xml(self, work_sheet):

        """Convert the tests from excel into xml for a given test work sheet"""

        # if isinstance(work_sheet, )


    def write_to_xml(self, xml_file):
        if os.path.exists(xml_file):
            logging.warning('The xml file will be over written: %s' % xml_file)


def callback_get_table_header():
    """Must return a tuple of Column headers"""
    return ('ID', 'Name', 'Description/Summary', 'Steps - Action', 'Steps - Expected',
            'Preconditions', 'Importance', 'Execution Type')


class TestLinkXMLXml2Excel(TestLinkXMLUtils):
    """Converts a given test link exported xml file into an excel file based on formatting options"""
    def __init__(self, xml_file, max_level=100):
        """

        :param max_level: the max_level of test suites
        :param xml_file: Source xml file
        """
        TestLinkXMLUtils.__init__(self)
        #### validation
        if not os.path.isfile(xml_file):
            raise IOError('The xml file does not exist: %s' % xml_file)

        #### parse the xml file
        xml_tree = xET.parse(xml_file)
        self.xml_root = xml_tree.getroot()

        self.load_testsuite_xml(max_level)

    def load_testsuite_xml(self, max_level):
        """

        :param max_level: the max_level of test suites
        :raise ValueError: For errors in work sheet names
        """

        #### input value processing
        if int(max_level) < 1:
            raise Exception('max_value must be integer value >= 1')

        self.suites, errs = self.get_suites(self.xml_root, max_level)

        self.suites[0]['name'] = ROOT_PLACEHOLDER
        suites_list = [item['name'].lower() for item in self.suites]
        for suite_name in suites_list:
            if suites_list.count(suite_name) > 1:
                errs.append(suite_name)
        if errs:
            raise ValueError("Duplicate suite names %s" % errs)

        #### extract top level suite names

    def write_to_excel(self, excel_file, skip_empty=False):
        """Writes the processed xml file into excel file

        :param excel_file: The name of output excel file
        :param skip_empty: Avoids creation of empty work sheets if no tests are present
        :return: None
        """

        if os.path.exists(excel_file):
            logging.warning('The excel file will be over written: %s' % excel_file)

        #### extract, process tests
        for item in self.suites:
            test_nodes = TestLinkXMLUtils.extract_tests(item['node'], item['leaf'])
            #copy, do not change the original dict
            #tests_dict = copy.deepcopy([TestLinkXMLUtils.process_testcase(test) for test in test_nodes])
            #Call this method if you want to prefix suite name
            tests_dict = copy.deepcopy([TestLinkXMLUtils.process_testcase(test, item['name']) for test in test_nodes])
            for test_d in tests_dict:
                self.callback_pre_clean_on_test_dict(test_d)
            item['tests'] = self.format_callback(tests_dict)
            # pprint.pprint(item['tests'])

        #### create worksheets and populate the tests
        workbook = xlsxwriter.Workbook(excel_file)
        format_text_wrap = workbook.add_format()
        format_text_wrap.set_text_wrap()
        format_detail_cells = workbook.add_format({
            'border': 1})
            #merge cells for detail
        for item in self.suites:
            if skip_empty and len(item['tests']) <= 0:
                continue
            sheet = workbook.add_worksheet(item['name'])
            sheet.merge_range('A1:H1', 'Detail', format_detail_cells)   # Merge the cells for detail
            sheet.set_row(0, ROW_HEIGHT*10)
            sheet.set_column('A:A', 5, format_text_wrap)
            sheet.set_column('B:B', 25, format_text_wrap)
            sheet.set_column('C:E', COL_WIDTH_STEPS, format_text_wrap)
            sheet.set_column('F:F', 25, format_text_wrap)
            sheet.set_column('G:H', 5, format_text_wrap)
            item['sheet'] = sheet
            for row, data in enumerate(item['tests']):
                data_l = hook_format_data_written_to_cell(data)
                sheet.write_row('A%s' % (row+HEAD_ROWS), data_l)

        # pprint.pprint(self.suites)
        workbook.close()

    def format_callback(self, list_of_test_dicts):
        """
        This will process a list of dicts obtained from process_testcase function. This must r
        :param list_of_test_dicts:
        :return: List of tuples
        """
        if type([]) != type(list_of_test_dicts):
            raise TypeError("list_of_test_dicts must of type list")
        for item in list_of_test_dicts:
            if type({}) != type(item):
                raise TypeError("All the elements in dictionary must be a dict")

        #header row
        ret_list = [callback_get_table_header()]

        #construct the csv file
        for test_d in list_of_test_dicts:
            ret_list.append(self.callback_get_test_tuple_from_test_d(test_d))
        return ret_list

    def callback_get_test_tuple_from_test_d(self, item_d):
        """Must return a tuple of Test contents
        :param item_d: Contains all required info about a test case.

        """
        steps_actions = steps_expected = custom_expected = ""
        if item_d['steps']:
            steps_actions = '\n'.join((step['actions'])
                                      if step['actions'] else "" for step in item_d['steps']).strip()
            steps_expected = '\n'.join((step['expectedresults'])
                                       if step['expectedresults'] else "" for step in item_d['steps']).strip()
        if item_d['custom_field']:
            custom_expected = '\n'.join((cus['value']) if cus['value'] else "" for cus in item_d['custom_field']
                                        if cus['name'] == EXPECTED_RESULTS).strip()

        if steps_expected == '' or custom_expected == '':
            expected_separator = ''
        else:
            expected_separator = '\n'

        # if not steps_actions:
        #     steps_actions = ''
        # if not item_d['summary']:
        #     item_d['summary'] = ''

        temp_list = [
            item_d['externalid'],
            item_d['name'],
            item_d['summary'],  # + '\n' + steps_actions,
            steps_actions,
            steps_expected + expected_separator + custom_expected,
            (item_d['preconditions']) if item_d['preconditions'] else '',
            item_d['importance'],
            # 'M' if str(item_d['execution_type']).strip() == '0' else 'A']
            str(item_d['execution_type']).strip()]
        return self.callback_post_clean_on_test_list(temp_list, (2, 3, 4, 5))

    def callback_post_clean_on_test_list(self, test_l, *args):
        """
        Write the required cleanup actions for a single test that will be executed after formatting is done
        :param test_l: Test data represented as a list
        :param args: A list/tuple which contains the list of indexes that must formatted
        :return:
        """
        if not args:
            raise IndexError('args must not be empty')
        if not isinstance(args[0], Iterable):
            raise TypeError('args[0] must a iterable')

        replace_regex_tuples = [('\n+\s+\n+', '\n'),
                                ('\n+\n+', '\n'),
                                ('Action - Step \d+$', ''),
                                ('Expected - Step \d+$', ''),
                                ]
        for index in args[0]:
            item = test_l[index]
            if not item or len(str(item).strip()) == 0:
                test_l[index] = ''
                continue

            item = html2text(item).strip()
            for regex_t in replace_regex_tuples:
                #pprint.pprint(re.findall(regex_t[0], item))
                item = re.sub(regex_t[0], regex_t[1], item)
                test_l[index] = item
        return tuple(test_l)

    def callback_pre_clean_on_test_dict(self, test_d):
        """
        Write the required cleanup actions for a single test before formatting
        :param test_d:
        """

        #note: write cleanup if required
        # deal with steps
        if 'steps' in test_d and isinstance(test_d['steps'], list) and test_d['steps']:
            new_field = []
            for item in test_d['steps']:
                new_field.append({key: item[key].replace('\n', '<br>\n') if isinstance(item[key], str) else item[key]
                                  for key in item})
            test_d['steps'] = new_field

        # deal with custom fields
        if 'custom_field' in test_d and isinstance(test_d['custom_field'], list) and test_d['custom_field']:
            new_field = []
            for item in test_d['custom_field']:
                new_field.append({key: item[key].replace('\n', '<br>\n') if isinstance(item[key], str) else item[key]
                                  for key in item})
            test_d['custom_field'] = new_field

        for item in test_d:
            # deal with strings
            if isinstance(test_d[item], str):
                test_d[item] = test_d[item].replace('\n', '<br>\n')

        return test_d


if __name__ == '__main__':
    input_xml_file = 'r:/all.xml'
    # input_xml_file = '/harvesters.xml'
    xlsx_file = 'r:/h.xlsx'
    x2xml = TestLinkXMLXml2Excel(input_xml_file, 10)
    x2xml.write_to_excel(xlsx_file)
    exit(100)



    ######## excel to xml ########
    # xml2xl = TestLinkXMLExcel2Xml('/test.xlsx')
    # xml2xl.write_to_xml('/op.xml')
